#!/usr/bin/env python3
"""Mirror low-frequency mobile cloud edits into SongBot's local query files.

This runs once when SongBot starts. Group message queries never call the cloud.
The desktop token stays in the local properties file and is never logged.
"""

from __future__ import annotations

import csv
import json
import os
import shutil
import sys
import tempfile
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SONGS_CSV = ROOT / "songs.csv"
STABLE_CSV = ROOT / "stable_info.csv"
STABLE_XLSX = ROOT.parent / "stable_info.xlsx"
CONFIG = ROOT / "data" / "cloud-announcement.properties"
STATE = ROOT / "data" / "mobile-library-sync-state.json"
MAX_RESPONSE = 4 * 1024 * 1024


def properties(path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            result[key.strip()] = value.strip()
    return result


def endpoint(value: str) -> str:
    parsed = urllib.parse.urlparse(value)
    if parsed.scheme != "https" or parsed.hostname not in {"editor.teacharm.moe", "bot-editor.vercel.app"}:
        raise RuntimeError("cloud API origin is not trusted")
    return urllib.parse.urlunparse(("https", parsed.netloc, "/api/mobile-data", "", "", ""))


def request_json(api: str, token: str, parameters: dict[str, str]) -> dict:
    url = f"{api}?{urllib.parse.urlencode(parameters)}"
    request = urllib.request.Request(url, headers={"Authorization": f"Desktop {token}", "Accept": "application/json"})
    with urllib.request.urlopen(request, timeout=30) as response:
        body = response.read(MAX_RESPONSE + 1)
    if len(body) > MAX_RESPONSE:
        raise RuntimeError("cloud response is too large")
    value = json.loads(body.decode("utf-8"))
    if not isinstance(value, dict):
        raise RuntimeError("cloud response is invalid")
    return value


def fetch_dataset(api: str, token: str, dataset: str) -> tuple[list[str], list[dict[str, str]], int]:
    offset = 0
    items: list[dict[str, str]] = []
    columns: list[str] = []
    revision = 0
    for _ in range(100):
        page = request_json(api, token, {"action": dataset, "offset": str(offset), "limit": "200"})
        if not columns:
            columns = [str(value) for value in page.get("columns", [])]
        revision = int(page.get("revision", 0))
        batch = page.get("items", [])
        if not isinstance(batch, list):
            raise RuntimeError(f"invalid {dataset} page")
        items.extend({str(key): "" if value is None else str(value) for key, value in row.items()} for row in batch)
        if not page.get("hasMore"):
            total = int(page.get("total", len(items)))
            if total != len(items) or not items or not columns:
                raise RuntimeError(f"incomplete {dataset} dataset")
            return columns, items, revision
        next_offset = int(page.get("nextOffset", offset + len(batch)))
        if not batch or next_offset <= offset:
            raise RuntimeError(f"invalid {dataset} pagination")
        offset = next_offset
    raise RuntimeError(f"too many {dataset} pages")


def atomic_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, path)


def write_csv(path: Path, columns: list[str], items: list[dict[str, str]]) -> Path:
    temporary = path.with_suffix(path.suffix + ".mobile-sync.tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=columns, extrasaction="ignore", lineterminator="\r\n")
        writer.writeheader()
        for item in items:
            writer.writerow({column: item.get(column, "") for column in columns})
    return temporary


def write_stable_workbook(columns: list[str], items: list[dict[str, str]]) -> Path | None:
    if not STABLE_XLSX.is_file():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    temporary = STABLE_XLSX.with_suffix(".mobile-sync.tmp.xlsx")
    workbook = load_workbook(STABLE_XLSX)
    sheet = workbook.worksheets[0]
    headers = [str(sheet.cell(1, index).value or "").strip() for index in range(1, sheet.max_column + 1)]
    if not headers or "sid" not in {value.lower() for value in headers}:
        workbook.close()
        raise RuntimeError("Stable workbook has no SID column")
    by_lower = {column.lower(): column for column in columns}
    for row_index, item in enumerate(items, start=2):
        for column_index, header in enumerate(headers, start=1):
            actual = by_lower.get(header.lower())
            if not actual:
                continue
            value = item.get(actual, "")
            if header.lower() == "cover" and value.upper() == "AUTO":
                directory = str(ROOT.parent / "stable_cover") + os.sep
                sheet.cell(row_index, column_index).value = f'="{directory}"&A{row_index}&".webp"'
            else:
                sheet.cell(row_index, column_index).value = value
    if sheet.max_row > len(items) + 1:
        sheet.delete_rows(len(items) + 2, sheet.max_row - len(items) - 1)
    workbook.save(temporary)
    workbook.close()
    return temporary


def main() -> int:
    if not CONFIG.is_file():
        return 0
    cfg = properties(CONFIG)
    token = cfg.get("desktopToken", "")
    if len(token) < 24:
        return 0
    api = endpoint(cfg.get("api", ""))
    status = request_json(api, token, {"action": "status"})
    cloud_revisions = {
        "songs": int((status.get("songs") or {}).get("revision", 0)),
        "stable": int((status.get("stable") or {}).get("revision", 0)),
    }
    previous = {}
    if STATE.is_file():
        try:
            previous = json.loads(STATE.read_text(encoding="utf-8"))
        except Exception:
            previous = {}
    if previous.get("revisions") == cloud_revisions:
        print("mobile cloud data is already synchronized")
        return 0

    song_columns, songs, song_revision = fetch_dataset(api, token, "songs")
    stable_columns, stable, stable_revision = fetch_dataset(api, token, "stable")
    if {"id", "song_name"} - {column.lower() for column in song_columns}:
        raise RuntimeError("cloud songs schema is incomplete")
    if {"sid", "title"} - {column.lower() for column in stable_columns}:
        raise RuntimeError("cloud Stable schema is incomplete")

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = ROOT / "backups" / "mobile-cloud-sync" / stamp
    backup.mkdir(parents=True, exist_ok=True)
    for source in (SONGS_CSV, STABLE_CSV, STABLE_XLSX):
        if source.is_file():
            shutil.copy2(source, backup / source.name)
    song_temp = write_csv(SONGS_CSV, song_columns, songs)
    stable_temp = write_csv(STABLE_CSV, stable_columns, stable)
    workbook_temp = write_stable_workbook(stable_columns, stable)
    try:
        os.replace(song_temp, SONGS_CSV)
        os.replace(stable_temp, STABLE_CSV)
        if workbook_temp is not None:
            os.replace(workbook_temp, STABLE_XLSX)
        atomic_json(STATE, {
            "schema": 1,
            "synchronizedAt": datetime.now().astimezone().isoformat(),
            "revisions": {"songs": song_revision, "stable": stable_revision},
            "rows": {"songs": len(songs), "stable": len(stable)},
        })
    finally:
        for temporary in (song_temp, stable_temp, workbook_temp):
            if temporary is not None:
                Path(temporary).unlink(missing_ok=True)
    print(f"mobile cloud synchronized: songs={len(songs)} stable={len(stable)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"mobile cloud sync skipped: {error}", file=sys.stderr)
        raise SystemExit(1)

